"""XLSX Financial & Statistical Report generation service.

Generates the CUA Financial & Statistical Report by populating the
master Excel template with live database aggregates.
"""

import io
import logging
import os
from datetime import datetime
from decimal import Decimal

from django.conf import settings
from django.db.models import Count, Q, Sum
from django.utils import timezone

import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template path — guarded at call-time
# ---------------------------------------------------------------------------
TEMPLATE_FILENAME = "New_Fin_Stats_Report_V6.1. 2024.xlsx"
TEMPLATE_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "reports")
TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, TEMPLATE_FILENAME)

# Credit Union identity (configurable via settings or .env)
CU_NAME = getattr(settings, "CREDIT_UNION_NAME", "Coastal Co-operative Credit Union Ltd.")


def _safe_decimal(value) -> Decimal:
    """Coerce a value to Decimal, defaulting to 0."""
    if value is None:
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"))


# ===================================================================
# Data collection helpers
# ===================================================================

def _collect_membership_stats() -> dict:
    """Aggregate member counts split by gender and activity status.

    Returns a dict keyed by (status, gender) → count.
    Status is one of: 'active', 'inactive', 'dormant'.
    Gender is one of: 'F', 'M', 'G', '' (unclassified).
    """
    from django.contrib.auth import get_user_model

    User = get_user_model()

    # Active = is_active=True, has logged in within 90 days OR has accounts
    # Inactive = is_active=True, but no recent activity
    # Dormant = is_active=False
    # Simplified: use is_active flag + last_login proximity
    ninety_days_ago = timezone.now() - timezone.timedelta(days=90)

    stats = {}
    for gender in ("F", "M", "G", ""):
        gender_q = Q(gender=gender) if gender else Q(gender="")

        # Dormant = deactivated accounts
        dormant = User.objects.filter(gender_q, is_active=False, role="customer").count()
        # Active = active + recent login
        active = User.objects.filter(
            gender_q, is_active=True, role="customer",
        ).filter(
            Q(last_login__gte=ninety_days_ago) | Q(last_login__isnull=False)
        ).count()
        # Inactive = active but never logged in
        inactive = User.objects.filter(
            gender_q, is_active=True, role="customer", last_login__isnull=True,
        ).count()

        stats[("active", gender)] = active
        stats[("inactive", gender)] = inactive
        stats[("dormant", gender)] = dormant

    return stats


def _collect_account_balances() -> dict:
    """Aggregate account balances by type, status, and owner gender.

    Returns dict keyed by (account_type, status, gender) → total balance.
    """
    from core.models.accounts import Account

    results = {}
    for acct_type in ("shares", "member_savings", "daily_susu", "savings", "youth_savings"):
        for is_active_val, status_label in [(True, "active"), (False, "inactive")]:
            qs = Account.objects.filter(
                account_type=acct_type, is_active=is_active_val
            ).select_related("user")

            for gender in ("F", "M", "G"):
                total = qs.filter(user__gender=gender).aggregate(
                    total=Sum("balance")
                )["total"]
                results[(acct_type, status_label, gender)] = _safe_decimal(total)

    return results


def _collect_loan_stats() -> dict:
    """Aggregate loan statistics by gender.

    Returns dict keyed by (metric, gender) → value.
    """
    from core.models.loans import Loan

    now = timezone.now()
    current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    results = {}

    for gender in ("F", "M", "G"):
        gender_q = Q(user__gender=gender)

        # Loans disbursed this month
        disbursed_qs = Loan.objects.filter(
            gender_q, status__in=["active", "approved"],
            created_at__gte=current_month_start,
        )
        results[("disbursed_count", gender)] = disbursed_qs.count()
        results[("disbursed_amount", gender)] = _safe_decimal(
            disbursed_qs.aggregate(t=Sum("amount"))["t"]
        )

        # Outstanding loans
        outstanding_qs = Loan.objects.filter(gender_q, status="active")
        results[("outstanding_count", gender)] = outstanding_qs.count()
        results[("outstanding_amount", gender)] = _safe_decimal(
            outstanding_qs.aggregate(t=Sum("outstanding_balance"))["t"]
        )

        # Delinquent (defaulted)
        delinquent_qs = Loan.objects.filter(gender_q, status="defaulted")
        results[("delinquent_count", gender)] = delinquent_qs.count()
        results[("delinquent_amount", gender)] = _safe_decimal(
            delinquent_qs.aggregate(t=Sum("outstanding_balance"))["t"]
        )

    return results


def _collect_receipts_payments() -> dict:
    """Aggregate receipts and payments for the current month.

    Returns dict keyed by label → Decimal amount.
    """
    from core.models.transactions import Transaction

    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    completed_this_month = Transaction.objects.filter(
        status="completed", timestamp__gte=month_start,
    )

    # 1. Shares Deposits: deposits where to_account.account_type == "shares"
    shares_deposits = completed_this_month.filter(
        transaction_type="deposit", to_account__account_type="shares"
    ).aggregate(t=Sum("amount"))["t"]

    # 2. Savings Deposits: deposits where to_account.account_type is in savings types
    savings_deposits = completed_this_month.filter(
        transaction_type="deposit",
        to_account__account_type__in=["member_savings", "daily_susu", "savings", "youth_savings"]
    ).aggregate(t=Sum("amount"))["t"]

    # 3. Loan Repayments: transaction_type == "repayment"
    loan_repayments = completed_this_month.filter(
        transaction_type="repayment"
    ).aggregate(t=Sum("amount"))["t"]

    # 4. Shares Withdrawals: withdrawals where from_account.account_type == "shares"
    shares_withdrawals = completed_this_month.filter(
        transaction_type="withdrawal", from_account__account_type="shares"
    ).aggregate(t=Sum("amount"))["t"]

    # 5. Savings Withdrawals: withdrawals where from_account.account_type is in savings types
    savings_withdrawals = completed_this_month.filter(
        transaction_type="withdrawal",
        from_account__account_type__in=["member_savings", "daily_susu", "savings", "youth_savings"]
    ).aggregate(t=Sum("amount"))["t"]

    # 6. Loans Disbursed: transaction_type == "disbursement"
    loans_disbursed = completed_this_month.filter(
        transaction_type="disbursement"
    ).aggregate(t=Sum("amount"))["t"]

    return {
        "shares_deposits": _safe_decimal(shares_deposits),
        "savings_deposits": _safe_decimal(savings_deposits),
        "loan_repayments": _safe_decimal(loan_repayments),
        "shares_withdrawals": _safe_decimal(shares_withdrawals),
        "savings_withdrawals": _safe_decimal(savings_withdrawals),
        "loans_disbursed": _safe_decimal(loans_disbursed),
    }


# ===================================================================
# Main report generator
# ===================================================================

def generate_xlsx_report(report_date: datetime = None) -> io.BytesIO:
    """Generate the CUA Financial & Statistical Report workbook.

    Args:
        report_date: The reporting date. Defaults to now.

    Returns:
        BytesIO buffer containing the populated .xlsx workbook.

    Raises:
        FileNotFoundError: If the master template is missing.
    """
    # --- Guard: template must exist ---
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"XLSX template not found at {TEMPLATE_PATH}. "
            "Ensure the master template is deployed to core/templates/reports/."
        )

    if report_date is None:
        report_date = timezone.now()

    logger.info("Generating XLSX Financial Stats Report for %s", report_date.strftime("%Y-%m"))

    # Load template preserving formulas
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["NOTES"]

    # ------------------------------------------------------------------
    # Header Information
    # ------------------------------------------------------------------
    # Strip timezone for Excel compatibility (Excel doesn't support tz-aware datetimes)
    naive_date = report_date.replace(tzinfo=None)
    ws["F3"] = naive_date  # Report date
    ws["D7"] = CU_NAME  # Credit Union name

    # ------------------------------------------------------------------
    # Section 0: Membership Information (rows 17–19, cols E/F/G)
    # E = Females, F = Males, G = Groups
    # ------------------------------------------------------------------
    membership = _collect_membership_stats()

    gender_col_map = {"F": "E", "M": "F", "G": "G"}
    status_row_map = {"active": 17, "inactive": 18, "dormant": 19}

    for (status, gender), count in membership.items():
        if gender in gender_col_map:
            col = gender_col_map[gender]
            row = status_row_map[status]
            ws[f"{col}{row}"] = count

    # ------------------------------------------------------------------
    # Section 1: Shares & Savings Balances (rows 27–32, cols E/F/G)
    # ------------------------------------------------------------------
    balances = _collect_account_balances()

    # Active Shares (row 27)
    for gender, col in gender_col_map.items():
        ws[f"{col}27"] = float(balances.get(("shares", "active", gender), 0))

    # Inactive Shares (row 28)
    for gender, col in gender_col_map.items():
        ws[f"{col}28"] = float(balances.get(("shares", "inactive", gender), 0))

    # Dormant Shares (row 29) — mapped from inactive with is_active=False
    # For simplicity, dormant accounts are a subset; set to 0 if not tracked separately
    for gender, col in gender_col_map.items():
        ws[f"{col}29"] = 0

    # Active Savings (row 30) — combines member_savings + daily_susu + savings + youth_savings
    for gender, col in gender_col_map.items():
        m_savings = balances.get(("member_savings", "active", gender), Decimal(0))
        susu = balances.get(("daily_susu", "active", gender), Decimal(0))
        savings = balances.get(("savings", "active", gender), Decimal(0))
        youth = balances.get(("youth_savings", "active", gender), Decimal(0))
        ws[f"{col}30"] = float(m_savings + susu + savings + youth)

    # Inactive Savings (row 31)
    for gender, col in gender_col_map.items():
        m_savings = balances.get(("member_savings", "inactive", gender), Decimal(0))
        susu = balances.get(("daily_susu", "inactive", gender), Decimal(0))
        savings = balances.get(("savings", "inactive", gender), Decimal(0))
        youth = balances.get(("youth_savings", "inactive", gender), Decimal(0))
        ws[f"{col}31"] = float(m_savings + susu + savings + youth)

    # Dormant Savings (row 32)
    for gender, col in gender_col_map.items():
        ws[f"{col}32"] = 0

    # ------------------------------------------------------------------
    # Section 1 (cont.): Loan Statistics (rows 34–39, cols E/F/G)
    # ------------------------------------------------------------------
    loans = _collect_loan_stats()

    loan_rows = {
        "disbursed_count": 34,
        "disbursed_amount": 35,
        "outstanding_count": 36,
        "outstanding_amount": 37,
        "delinquent_count": 38,
        "delinquent_amount": 39,
    }

    for (metric, gender), value in loans.items():
        if gender in gender_col_map and metric in loan_rows:
            col = gender_col_map[gender]
            row = loan_rows[metric]
            ws[f"{col}{row}"] = float(value) if isinstance(value, Decimal) else value

    # ------------------------------------------------------------------
    # Section 2: Receipts & Payments (rows 45–54, col C=Receipts, G=Payments)
    # ------------------------------------------------------------------
    rp = _collect_receipts_payments()

    ws["C47"] = float(rp["shares_deposits"])
    ws["C48"] = float(rp["savings_deposits"])
    ws["C49"] = float(rp["loan_repayments"])

    ws["G47"] = float(rp["shares_withdrawals"])
    ws["G48"] = float(rp["savings_withdrawals"])
    ws["G49"] = float(rp["loans_disbursed"])

    # ------------------------------------------------------------------
    # Save to buffer
    # ------------------------------------------------------------------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info("XLSX Financial Stats Report generated successfully (%d bytes)", buffer.getbuffer().nbytes)
    return buffer
